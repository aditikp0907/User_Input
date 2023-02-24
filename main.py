import tkinter as tk
from tkinter.simpledialog import askinteger, askstring
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import datetime

DEFAULT_COLOR = "#DCDCDC"

def change_color(box_num, color, box_name=None):
    if color == "R":
        boxes[box_num - 1]["bg"] = "red"
    elif color == "G":
        boxes[box_num - 1]["bg"] = "green"
    elif color == "B":
        boxes[box_num - 1]["bg"] = "blue"
    elif color == "Y":
        boxes[box_num - 1]["bg"] = "yellow"
    elif color == "Z":
        boxes[box_num - 1]["bg"] = DEFAULT_COLOR

    if box_name is not None:
        if box_name.lower() == "reset":
            box_name = f"Box {box_num}"
        boxes[box_num - 1]["text"] = box_name

    # write to the log worksheet
    log_ws.append([box_num, color, box_name, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    log_ws.column_dimensions[get_column_letter(1)].width = 10
    log_ws.column_dimensions[get_column_letter(2)].width = 10
    log_ws.column_dimensions[get_column_letter(3)].width = 20
    log_ws.column_dimensions[get_column_letter(4)].width = 20

root = tk.Tk()
root.attributes("-fullscreen", True)
# Prompt the user for the number of rows and columns
num_rows = askinteger("Number of Rows", "Enter the number of rows:")
num_cols = askinteger("Number of Columns", "Enter the number of columns:")

# create a new workbook and worksheet for logging
wb = Workbook()
log_ws = wb.active
log_ws.title = "Log"
log_ws.append(["Box Number", "Color", "Box Name", "Time Stamp"])
log_ws['A1'].font = Font(bold=True)
log_ws['B1'].font = Font(bold=True)
log_ws['C1'].font = Font(bold=True)
log_ws['D1'].font = Font(bold=True)
log_ws.row_dimensions[1].height = 20
for cell in log_ws[1]:
    cell.alignment = Alignment(horizontal='center', vertical='center')

boxes = []

for i in range(num_rows * num_cols):
    box = tk.Label(root, text=f"Box {i+1}", bg=DEFAULT_COLOR, width=15, height=5, font=("Arial", 20), bd=2, relief="solid", justify="center")
    box.grid(row=i // num_cols, column=i % num_cols, padx=10, pady=10)
    boxes.append(box)

# Center the boxes
for i in range(num_cols):
    root.grid_columnconfigure(i, weight=1)
for i in range(num_rows):
    root.grid_rowconfigure(i, weight=1)

# Center the window
x = (root.winfo_screenwidth() - root.winfo_reqwidth()) / 2
y = (root.winfo_screenheight() - root.winfo_reqheight()) / 2
root.geometry("+%d+%d" % (x, y))

while True:
    try:
        user_input = askstring("Box Input",
                               "Enter box number, color, and optional box name (e.g. 1R, Box 1) or press Cancel to quit:",
                               parent=root)
        if not user_input:
            break
        box_num, color = user_input[:2]
        box_num = int(box_num)
        color = color.upper()
        if len(user_input) > 2:
            box_name = user_input[3:]
        else:
            box_name = None
        change_color(box_num, color, box_name)

        # write data to the Excel file
        wb.save('D:/Users/Aditi/Onedrive/Desktop/CodeIntern/Code intern.xlsx')

        root.update()

        # Display a statement on the console
        color_name = ""
        if color == "R":
            color_name = "red"
        elif color == "G":
            color_name = "green"
        elif color == "B":
            color_name = "blue"
        elif color == "Y":
            color_name = "yellow"
        elif color == "Z":
            color_name = "default"

        box_label = boxes[box_num - 1]["text"] if boxes[box_num - 1]["text"] else f"Box {box_num}"
        print(
            f"The box selected is box number {box_num}. The color selected is {color_name}. And the name of the box is {box_label}.")

    except (ValueError, IndexError):
        print(
            "Invalid input. Please enter a valid box number (1-{}), followed by a valid color (R, G, B, Y, Z), and an optional box name.".format(
                len(boxes)))

root.mainloop()

